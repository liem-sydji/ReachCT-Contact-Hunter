"""
ReachCT — reachct.py
Main scraper. Searches Google Maps and extracts company contact info.

Requirements:
    pip install playwright openpyxl beautifulsoup4 spacy nltk
    playwright install chromium
    python -m spacy download es_core_news_sm
    python -c "import nltk; nltk.download('stopwords')"

Usage:
    python reachct.py --query "agencia de marketing" --city "Madrid" --country "España" --start 0 --end 25
    python reachct.py --export --city "Madrid" --country "España"
"""

import re
import csv
import uuid
import random
import argparse
import asyncio
from urllib.parse import urlparse
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from database import init_db, save_search, upsert_company, get_companies
from verification import verify

# ── CONFIG ────────────────────────────────────────────────────────────────────
HEADLESS = True

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
]

CONTACT_PATHS = [
    "/contact", "/contact-us", "/contacto", "/contactanos",
    "/about", "/about-us", "/sobre-nosotros", "/info", "/equipo"
]

EMAIL_BLACKLIST = [
    "example", "domain", "youremail", "user@", "email@",
    "sentry", "wix", "wordpress", "jquery", "schema",
    "png", "jpg", "gif", "svg", "css", "js@", "noreply",
]

MAX_RETRIES = 0  # 0 retries = 1 attempt only, fail fast


# ─────────────────────────────────────────────────────────────────────────────


# ── Helpers ───────────────────────────────────────────────────────────────────

def random_delay(min_ms=2000, max_ms=5000):
    return random.randint(min_ms, max_ms)


def extract_emails(text: str) -> list:
    found = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    return list(dict.fromkeys([
        e.lower() for e in found
        if not any(b in e.lower() for b in EMAIL_BLACKLIST)
    ]))


def extract_phones(text: str) -> list:
    found = re.findall(r"(?:\+?\d[\s\-.]?){7,15}", text)
    return list(dict.fromkeys([
        p.strip() for p in found
        if 7 <= len(re.sub(r"\D", "", p)) <= 15
    ]))


# ── Website scraping ──────────────────────────────────────────────────────────

async def try_contact_pages(page, base_url: str) -> dict:
    parsed = urlparse(base_url)
    base = f"{parsed.scheme}://{parsed.netloc}"
    for path in CONTACT_PATHS:
        try:
            r = await page.goto(base + path, timeout=8000, wait_until="domcontentloaded")
            if r and r.status == 200:
                await page.wait_for_timeout(random_delay(800, 1500))
                html = await page.content()
                text = await page.evaluate("() => document.body.innerText")
                emails = extract_emails(html + " " + text)
                if emails:
                    phones = extract_phones(html + " " + text)
                    return {"emails": emails, "phones": phones}
        except:
            continue
    return {"emails": [], "phones": []}


async def scrape_website(browser, url: str, retries: int = MAX_RETRIES) -> dict:
    if not url:
        return {"email": "", "phone": "", "page_text": ""}

    # Skip HTTP sites — usually old/slow/dead, not worth waiting 15s
    if url.startswith("http://"):
        print(f"    ⏭️  Skipped HTTP site: {url[:50]}")
        return {"email": "", "phone": "", "page_text": ""}

    for attempt in range(retries):
        context = None
        try:
            context = await browser.new_context(
                user_agent=random.choice(USER_AGENTS)
            )
            page = await context.new_page()
            email = ""
            phone = ""
            page_text = ""

            await page.goto(url, timeout=8000, wait_until="domcontentloaded")
            await page.wait_for_timeout(random_delay(300, 600))
            html = await page.content()
            page_text = await page.evaluate("() => document.body.innerText")

            emails = extract_emails(html + " " + page_text)
            phones = extract_phones(html + " " + page_text)

            if emails:
                email = emails[0]
            if phones:
                phone = phones[0]

            if not email:
                data = await try_contact_pages(page, url)
                if data["emails"]:
                    email = data["emails"][0]
                if not phone and data["phones"]:
                    phone = data["phones"][0]

            return {"email": email, "phone": phone, "page_text": page_text}

        except Exception as e:
            print(f"    ⚠️  Skipped (timeout): {url[:50]}")
        finally:
            # Always close context to free RAM
            if context:
                try:
                    await context.close()
                except:
                    pass

    return {"email": "", "phone": "", "page_text": ""}


# ── Name extraction ───────────────────────────────────────────────────────────

async def get_business_name(page) -> str:
    SKIP = {"resultados", "results", "google maps", "google", "", "maps"}

    for panel_sel in ['div[role="main"]', 'div.m6QErb']:
        try:
            await page.wait_for_selector(panel_sel, timeout=4000)
            break
        except:
            continue

    await page.wait_for_timeout(random_delay(800, 1200))

    for sel in ['h1.DUwDvf', 'div.DUwDvf', 'h1.fontHeadlineLarge',
                'div.fontHeadlineLarge', 'div[role="main"] h1']:
        try:
            els = await page.locator(sel).all()
            for el in els:
                try:
                    candidate = (await el.inner_text(timeout=2000)).strip()
                    if candidate.lower() not in SKIP and len(candidate) > 1:
                        return candidate
                except:
                    continue
        except:
            continue

    try:
        slug = page.url.split("/maps/place/")[1].split("/")[0]
        name = slug.replace("+", " ").replace("%20", " ").strip()
        if name.lower() not in SKIP:
            return name
    except:
        pass

    return ""


# ── Google Maps scraper ───────────────────────────────────────────────────────

async def scrape_google_maps(query: str, city: str, country: str,
                             start_idx: int, end_idx: int, run_id: str,
                             jobs: dict = None, job_id: str = None) -> list:
    results = []
    location = f"{city}, {country}"
    search = f"{query} {location}"
    maps_url = f"https://www.google.com/maps/search/{search.replace(' ', '+')}"

    print(f"\n🗺️  ReachCT — Searching: '{search}'")
    print(f"   Range: {start_idx} → {end_idx}")
    print(f"   {maps_url}\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            locale="es-ES",
            viewport={"width": 1280, "height": 800},
        )
        page = await context.new_page()
        await page.goto(maps_url, timeout=30000)
        await page.wait_for_timeout(random_delay(1500, 2000))

        # Accept cookies
        try:
            for btn_text in ["Aceptar todo", "Accept all", "Aceptar", "Accept"]:
                btn = page.get_by_role("button", name=btn_text)
                if await btn.count() > 0:
                    await btn.first.click()
                    print("🍪 Accepted cookies")
                    await page.wait_for_timeout(random_delay(1000, 1500))
                    break
        except:
            pass

        print(f"📄 {await page.title()}")

        # Scroll in chunks of 25 — only load what we need
        # For start=100, end=125 we only need to scroll to 125, not load all 125 at once
        print(f"📜 Scrolling to load listings {start_idx}→{end_idx}...")
        seen_hrefs = set()
        all_hrefs = []
        scroll_count = 0
        max_scrolls = (end_idx // 3) + 6

        while len(all_hrefs) < end_idx and scroll_count < max_scrolls:
            # Collect current visible HREFs
            links = await page.locator('a[href*="/maps/place/"]').all()
            for l in links:
                try:
                    href = await l.get_attribute("href") or ""
                    if "/maps/place/" in href and href not in seen_hrefs:
                        seen_hrefs.add(href)
                        all_hrefs.append(href)
                except:
                    continue

            # Stop scrolling if we have enough
            if len(all_hrefs) >= end_idx:
                print(f"   ✅ Loaded {len(all_hrefs)} listings — stopping scroll")
                break

            # Scroll more
            try:
                for sel in ['div[role="feed"]', 'div[aria-label*="Resultados"]', 'div[aria-label*="Results"]']:
                    feed = page.locator(sel)
                    if await feed.count() > 0:
                        await feed.evaluate("el => el.scrollTop += 1500")
                        break
                await page.wait_for_timeout(random_delay(600, 1000))
            except:
                break
            scroll_count += 1

        batch = all_hrefs[start_idx:end_idx]
        total = len(batch)
        total_on_maps = len(all_hrefs)

        # Update job with total listings found on Maps
        if jobs and job_id:
            jobs[job_id]["total_on_maps"] = total_on_maps
            jobs[job_id]["processing"] = total
        print(f"\n✅ {len(all_hrefs)} total listings — processing {start_idx}→{start_idx + total}\n")

        if total == 0:
            print("❌ No listings in range. Try scrolling further or adjusting --start/--end.")
            await page.wait_for_timeout(4000)
            await browser.close()
            return []

        # batch is already HREFs (strings) — no need to extract
        batch_hrefs = batch

        BROWSER_RESTART_EVERY = 10

        for i, href in enumerate(batch_hrefs):
            # Check for cancellation between listings
            if jobs and job_id and jobs.get(job_id, {}).get("status") == "cancelling":
                print(f"  🛑 Search cancelled at listing {start_idx + i + 1} — saving {len(results)} results")
                break

            # Restart browser every 10 listings to free RAM
            if i > 0 and i % BROWSER_RESTART_EVERY == 0:
                print(f"  🔄 Restarting browser (listing {start_idx + i + 1})...")
                try:
                    await browser.close()
                except:
                    pass
                browser = await p.chromium.launch(headless=HEADLESS)
                context = await browser.new_context(
                    user_agent=random.choice(USER_AGENTS),
                    locale="es-ES",
                    viewport={"width": 1280, "height": 800},
                )
                page = await context.new_page()
                print(f"  ✅ Browser restarted cleanly")

            try:
                # Navigate directly to the listing URL instead of clicking stale element
                await page.goto(href, timeout=15000, wait_until="domcontentloaded")
                await page.wait_for_timeout(random_delay(1500, 2500))

                name = await get_business_name(page)

                # Phone from Maps
                phone_maps = ""
                try:
                    btn = page.locator('button[data-item-id*="phone"]')
                    if await btn.count() > 0:
                        raw = await btn.first.get_attribute("data-item-id") or ""
                        phone_maps = raw.replace("phone:tel:", "").strip()
                except:
                    pass

                # Website
                website = ""
                for web_sel in ['a[data-item-id="authority"]',
                                'a[aria-label*="sitio web" i]',
                                'a[aria-label*="website" i]']:
                    try:
                        el = page.locator(web_sel).first
                        if await el.count() > 0:
                            href = await el.get_attribute("href") or ""
                            if href and "google" not in href:
                                website = href
                                break
                    except:
                        continue

                # Address
                address = ""
                try:
                    addr = page.locator('button[data-item-id*="address"]')
                    if await addr.count() > 0:
                        address = await addr.first.inner_text(timeout=2000)
                except:
                    pass

                print(f"[{start_idx + i + 1}] {name or '(no name)'}")
                print(f"  📞 {phone_maps or 'no phone'}")
                print(f"  🌐 {website[:65] if website else 'no website'}")

                # Scrape website for email if URL available
                web_data = {"email": "", "phone": "", "page_text": ""}
                if website:
                    web_data = await scrape_website(browser, website)

                # Verify
                v = verify(web_data["page_text"], website, query)

                print(f"  🔎 {v['category']} — {v['reason']}")

                final_phone = phone_maps or web_data["phone"]
                email = web_data["email"]
                print(f"  ✉️  {email or 'no email'}")

                results.append({
                    "run_id": run_id,
                    "name": name,
                    "email": email,
                    "phone": final_phone,
                    "website": website,
                    "city": city,
                    "country": country,
                    "company_type": query,
                    "category": v["category"],
                    "maps_url": page.url,
                })
                print()

                # Random delay between listings to avoid detection
                await page.wait_for_timeout(random_delay(300, 600))

            except Exception as e:
                print(f"  ⚠️  Error on listing {start_idx + i + 1}: {e}\n")
                continue

        await browser.close()

    return results


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(data: list, query: str, city: str, country: str) -> str:
    if not data:
        print("⚠️  No data to export.")
        return ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Styles
    hdr_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    hdr_font = Font(name="Arial", color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    wht_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ttl_fill = PatternFill(start_color="D6DCFF", end_color="D6DCFF", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    d_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="C5C5C5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Company Name", "Email", "Phone Number", "Website", "Location", "Company Type", "Status"]
    col_widths = [32, 35, 20, 40, 22, 25, 18]
    num_cols = len(headers)

    # Title
    ws.merge_cells(f"A1:{chr(64 + num_cols)}1")
    t = ws["A1"]
    t.value = f"ReachCT  |  {query}  |  {city}, {country}  |  {datetime.now().strftime('%Y-%m-%d')}"
    t.font = Font(name="Arial", bold=True, size=12, color="1A1A2E")
    t.alignment = center
    t.fill = ttl_fill
    ws.row_dimensions[1].height = 30

    # Headers
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 26

    # Data
    for row_i, item in enumerate(data, 3):
        fill = alt_fill if row_i % 2 == 0 else wht_fill
        values = [
            item.get("name", ""),
            item.get("email", ""),
            item.get("phone", ""),
            item.get("website", ""),
            f"{item.get('city', '')}, {item.get('country', '')}",
            item.get("company_type", ""),
            item.get("category", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill = fill
            cell.font = d_font
            cell.border = border
            cell.alignment = left
        ws.row_dimensions[row_i].height = 20

    # Footer
    fr = len(data) + 3
    ws.merge_cells(f"A{fr}:{chr(64 + num_cols)}{fr}")
    f = ws[f"A{fr}"]
    f.value = (
        f"Total: {len(data)}   |   "
        f"Emails: {sum(1 for r in data if r.get('email'))}   |   "
        f"Phones: {sum(1 for r in data if r.get('phone'))}   |   "
        f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    f.font = Font(name="Arial", size=9, italic=True, color="888888")
    f.alignment = center
    ws.row_dimensions[fr].height = 18
    ws.freeze_panes = "A3"

    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    loc_slug = f"{city}_{country}".replace(" ", "_").replace(",", "")
    filename = f"reachct_{query.replace(' ', '_')}_{loc_slug}_{stamp}.xlsx"
    wb.save(filename)
    print(f"\n📊 Excel saved: {filename}")
    return filename


# ── Entry point ───────────────────────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(description="ReachCT — Google Maps Contact Scraper")
    parser.add_argument("--query", default="empresa de software", help="Type of business")
    parser.add_argument("--city", default="Madrid", help="City to search in")
    parser.add_argument("--country", default="España", help="Country to search in")
    parser.add_argument("--start", default=0, type=int, help="Start index in Maps listing")
    parser.add_argument("--end", default=25, type=int, help="End index in Maps listing")
    parser.add_argument("--export", action="store_true", help="Export existing DB data to Excel without scraping")
    args = parser.parse_args()

    init_db()

    # ── Export mode ───────────────────────────────────────────────────────────
    if args.export:
        print(f"\n📦 Exporting from database: {args.city}, {args.country}")
        data = get_companies(city=args.city, country=args.country)
        if data:
            export_to_excel(data, args.query, args.city, args.country)
            print(f"✅ Exported {len(data)} companies.")
        else:
            print("⚠️  No companies found in DB for that location.")
        return

    # ── Scrape mode ───────────────────────────────────────────────────────────
    run_id = str(uuid.uuid4())[:8]
    print("=" * 55)
    print("  🔍  ReachCT")
    print("=" * 55)
    print(f"  Query   : {args.query}")
    print(f"  City    : {args.city}")
    print(f"  Country : {args.country}")
    print(f"  Range   : {args.start} → {args.end}")
    print(f"  Run ID  : {run_id}")
    print("=" * 55)

    # Clean inputs — remove trailing spaces and normalise case
    clean_query = args.query.strip()
    clean_city = args.city.strip().title()
    clean_country = args.country.strip().title()

    results = await scrape_google_maps(
        clean_query, clean_city, clean_country,
        args.start, args.end, run_id
    )

    if not results:
        print("\n❌ No verified companies found.")
        return

    # Save to DB
    inserted = updated = skipped = 0
    for company in results:
        status = upsert_company(run_id, company)
        if status == "inserted":
            inserted += 1
        elif status == "updated":
            updated += 1
        else:
            skipped += 1

    save_search(run_id, args.query, args.city, args.country,
                args.start, args.end, len(results))

    # Export Excel
    filename = export_to_excel(results, args.query, args.city, args.country)

    print(f"\n📊 Summary")
    print(f"   Scraped  : {len(results)}")
    print(f"   Inserted : {inserted}")
    print(f"   Updated  : {updated}")
    print(f"   Skipped  : {skipped}")
    print(f"\n✅ Done! Open '{filename}'")


if __name__ == "__main__":
    asyncio.run(main())
